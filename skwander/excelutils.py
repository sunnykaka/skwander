# -*- coding: utf-8 -*-

import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.cell import get_column_letter


def write_designer_excel(designer, file_path, filename):

    excel_path = '%s.xlsx' % file_path
    wb = Workbook()
    ws1 = wb.active

    # designer info
    designer_header_row = 1
    designer_content_row = 2
    designer_name_col = 3
    designer_headers = [{'name': u"序号", 'width': 6}, {'name': u"设计师ID", 'width': 12},
                        {'name': u"设计师名称", 'width': 40}, {'name': u"国籍", 'width': 15},
                        {'name': u"描述", 'width': 15},
                        {'name': u"图片名称", 'width': 12},
                        {'name': u"图片链接", 'width': 12, 'hyperlink': True},
                        {'name': u"设计师链接", 'width': 20, 'hyperlink': True}]

    write_headers(ws1, designer_header_row, designer_headers)

    designer_cols = [1, designer['uid'],
                     designer['name'], designer.get('nation', ''),
                     designer['desc'],
                     designer['img_names'],
                     designer.get('img_url', ''),
                     designer['url']
                     ] if 'desc' in designer else []

    write_content_rows(ws1, designer_cols, designer_headers, designer_content_row, designer_name_col)

    # product info
    if designer['products']:
        product_header_row = designer_content_row + 2
        product_content_row = product_header_row + 1
        product_name_col = 3
        product_headers = [{'name': u"序号", 'width': 6}, {'name': u"商品ID", 'width': 12},
                           {'name': u"商品名称", 'width': 40},
                           {'name': u"当前价格", 'width': 15}, {'name': u"原价", 'width': 15},
                           {'name': u"尺码", 'width': 12}, {'name': u"库存", 'width': 12},
                           {'name': u"可选尺码信息", 'width': 20}, {'name': u"尺码表", 'width': 30},
                           {'name': u"商品描述", 'width': 80},
                           {'name': u"图片名称", 'width': 20},
                           {'name': u"图片链接", 'width': 20},
                           {'name': u"商品链接", 'width': 20, 'hyperlink': True}]

        write_headers(ws1, product_header_row, product_headers)

        for i, p in enumerate(designer['products']):
            product_cols = [i + 1, p['uid'],
                            p['name'],
                            p.get('price', ''), p.get('original_price', ''),
                            p.get('current_size', ''), p.get('stock', ''),
                            p.show_size_info(), p.show_design_size(),
                            p.get('desc', ''),
                            '\n'.join(p['img_names']),
                            '\n'.join(p['img_url']),
                            p.show_url()]

            row = i + product_content_row

            write_content_rows(ws1, product_cols, product_headers, row, product_name_col)

    wb.save(filename=excel_path)


def write_content_rows(ws, cols, headers, row, name_col):
    for i, value in enumerate(cols):
        col = i + 1
        cell = ws.cell(column=col, row=row, value=value)
        if headers[i].get('hyperlink'):
            cell.hyperlink = value
        # if col == name_col:
        #     cell.hyperlink = cols[-1]


def write_headers(ws, row, headers):
    for i, h in enumerate(headers):
        col = i + 1
        cell = ws.cell(column=col, row=row, value=h['name'])
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col)].width = h['width']



