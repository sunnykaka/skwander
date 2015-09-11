# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, fills, Color


wb = Workbook()

dest_filename = '/tmp/empty_book.xlsx'

ws1 = wb.active
ws1.title = "range names"

for row in range(1, 40):
    ws1.append(range(600))

ws2 = wb.create_sheet(title="Pi")

ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        x = ws3.cell(column=col, row=row, value=u"%s 你好" % get_column_letter(col))
        x.font = Font(bold=True)
        x.alignment = Alignment(horizontal='center')
        x.fill = PatternFill(patternType=fills.FILL_SOLID, fgColor=Color('FF869702'))
        ws3.column_dimensions[get_column_letter(col)].width = 20
print(ws3['AA10'].value)

wb.save(filename=dest_filename)
